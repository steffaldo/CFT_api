import streamlit as st
import pandas as pd
import os
from openpyxl import load_workbook
from config.config_loader import load_toml
from components.data_cleaning import *
import streamlit_notify as stn
from utils.api_parser import submit_new_surveys
from data.supabase import (
    get_dairy_inputs,
    upsert_dairy_inputs,
    upsert_outputs_from_df,
)
import unicodedata
import re


st.title("CFT Dairy Data Upload")

# Add debug tick


if st.sidebar.checkbox("Debug mode (show extra info)", value=st.session_state.debug):
    st.session_state.debug = True

st.markdown(
    "Upload CFT dairy survey workbooks, check for data quality issues, "
    "and send them to the CFT model for emissions results."
)

st.divider()

st.info(
    "**Workflow:** 1) Upload survey workbook(s) • 2) Fix any data issues "
    "• 3) Upload to database & run CFT API."
)

stn.notify()

data = get_dairy_inputs()
df = pd.DataFrame(data)

herd_sections = load_toml("herd.toml")["herd_section"]
herd_varieties = load_toml("herd.toml")["herd_variety"]


# ---- Drop files
if "uploader_key" not in st.session_state:
    st.session_state.uploader_key = 0

survey_dump = st.file_uploader(
    "Upload new survey workbook(s)",
    accept_multiple_files=True,
    type="xlsx",
    help=(
        "Upload one or more Excel (.xlsx) CFT survey templates. "
        "Each file should contain data for a single farm and milk year."
    ),
    key=f"uploader{st.session_state.uploader_key}",
)

# initialise mapping schema api
schema_path = os.path.join("schema", "input_schema_mapping.csv")
input_schema = pd.read_csv(schema_path)

# Build structured mapping
schema_dict = {
    row.metric: {
        "cell": row.survey_mapping,
        "type": row.types,
        "default": row.default_value if "default_value" in row else None
    }
    for _, row in input_schema.iterrows()
}

# Extract columns
input_columns = list(schema_dict.keys())

# Init empty df to hold ingested data
survey_loader = pd.DataFrame()

feed_conversion_mapping = {
    "fwi_select": "C61",
    "dmi_select": "D61",
    "feed_per_animal": "C64",
    "feed_per_herd": "D64",
    "feed_period_day_single": "C67",
    "feed_period_day_custom": "D67"
}

# feed lookup dict
feed_items = load_toml("feed.toml")["feed"]
feed_meta = {
    f["cft_name"]: f
    for f in feed_items
}

# Cleaner function for invisible values in cells
def cell_has_value(cell):
    if cell is None:
        return False
    if pd.isna(cell):
        return False
    if isinstance(cell, str):
        return cell.strip() != ""
    return True

# feed normalisation function
def normalize_feed_value(
    *,
    value,
    metric,
    feed_meta,
    row_data,
    dmi_conversion,
    herd_feed_indicator,
    multiday_feed_indicator,
    debug=False,
):
    """
    Normalizes feed values to kgDMI_head_day.

    Expects metric format:
    feed.<feed_name>.<hs_name>.kgDMI_head_day
    """

    # ---- parse metric ----
    try:
        _, feed_name, hs_name, _ = metric.split(".", 3)
    except ValueError:
        st.error(f"Invalid feed metric format: {metric}")
        st.stop()

    feed_info = feed_meta.get(feed_name)
    if feed_info is None:
        st.error(f"Unknown feed type in column name: {feed_name}")
        st.stop()


    # ---- 1. FWI → DMI ----
    if dmi_conversion:
        conversion_factor = feed_info.get("fwi_to_dmi")

        if debug:
            st.write(f"FWI→DMI factor for {feed_name}:", conversion_factor)

        if conversion_factor is None:
            st.error(f"Missing FWI→DMI conversion factor for feed: {feed_name}")
            st.stop()

        value = value * conversion_factor

        if debug:
            st.write(f"After FWI→DMI ({feed_name}):", value)

    # ---- 2. herd → head ----
    if herd_feed_indicator:
        herd_count_key = f"{hs_name}.herd_count"
        herd_count = row_data.get(herd_count_key)

        if debug:
            st.write(f"Herd count for {hs_name}:", herd_count)

        if herd_count is None or herd_count <= 0:
            st.error(
                f"Herd count missing or invalid for herd type '{hs_name}' "
                f"(expected column '{herd_count_key}')"
            )
            st.stop()

        value = value / herd_count

        if debug:
            st.write(f"After herd→head ({hs_name}):", value)

    # ---- 3. multi-day → per-day ----
    if multiday_feed_indicator > 1:
        value = value / multiday_feed_indicator

        if debug:
            st.write(
                f"After multi-day→day (/{multiday_feed_indicator}):",
                value
            )

    return round(value, 6) if value is not None else None

# text slugify function for farm names
def slugify(text: str) -> str:
    # Normalize accented characters → ASCII
    text = unicodedata.normalize("NFKD", text)
    text = text.encode("ascii", "ignore").decode("ascii")

    # Lowercase
    text = text.lower()

    # Replace non-alphanumeric with hyphens
    text = re.sub(r"[^a-z0-9]+", "-", text)

    # Trim hyphens from start/end
    text = text.strip("-")

    return text


# Loop and ingest each survey
for survey in survey_dump:

    wb = load_workbook(survey, data_only=True)
    ws = wb.active

    row_data = {}
    skip_survey = False

    # helper
    def fail(msg):
        st.error(f"❌ {survey.name} skipped — {msg}")
        return True

    # =====================================================
    # FEED CONFIG VALIDATION (RUN ONCE PER FILE)
    # =====================================================

    # --- DMI vs FWI ---
    dmi_selected = cell_has_value(ws[feed_conversion_mapping["dmi_select"]].value)
    fwi_selected = cell_has_value(ws[feed_conversion_mapping["fwi_select"]].value)

    if dmi_selected and fwi_selected:
        skip_survey = fail("Both DMI and FWI selected")
    elif not dmi_selected and not fwi_selected:
        skip_survey = fail("No feed unit selected (DMI/FWI)")
    else:
        dmi_conversion = fwi_selected

    # --- Per animal vs herd ---
    animal_selected = cell_has_value(ws[feed_conversion_mapping["feed_per_animal"]].value)
    herd_selected = cell_has_value(ws[feed_conversion_mapping["feed_per_herd"]].value)

    if animal_selected and herd_selected:
        skip_survey = fail("Both per-animal and per-herd feed selected")
    elif not animal_selected and not herd_selected:
        skip_survey = fail("Missing animal/herd feed selection")
    else:
        herd_feed_indicator = herd_selected

    # --- Feeding period ---
    day_selected = cell_has_value(ws[feed_conversion_mapping["feed_period_day_single"]].value)
    custom_day_selected = cell_has_value(ws[feed_conversion_mapping["feed_period_day_custom"]].value)

    if day_selected and custom_day_selected:
        skip_survey = fail("Both single-day and multi-day feeding selected")
    elif not day_selected and not custom_day_selected:
        skip_survey = fail("Missing feeding period selection")
    elif day_selected:
        multiday_feed_indicator = 1
    else:
        multiday_feed_indicator = ws[feed_conversion_mapping["feed_period_day_custom"]].value

    if skip_survey:
        continue


    # ---- Hard Checkpoint: farm_id must exist ----
    farm_id_cell = schema_dict["farm_id"]["cell"]
    raw_farm_id = ws[farm_id_cell].value
    if not cell_has_value(raw_farm_id):
        st.error(f"❌ {survey.name} skipped — missing Farm Name ({farm_id_cell})")
        continue

    # ---- Hard Checkpoint: milk_year must exist ----
    milk_year_cell = schema_dict["milk_year"]["cell"]
    raw_milk_year = ws[milk_year_cell].value
    if not cell_has_value(raw_milk_year):
        st.error(f"❌ {survey.name} skipped — missing milk_year ({milk_year_cell})")
        continue


    # =====================================================
    # METRIC EXTRACTION LOOP
    # =====================================================
    for metric, info in schema_dict.items():
        try:
            value = ws[info["cell"]].value

            if not cell_has_value(value) and cell_has_value(info["default"]):
                value = info["default"]

            if info["type"] == "int":
                value = int(value) if cell_has_value(value) else None
            elif info["type"] == "float":
                value = round(float(value), 6) if cell_has_value(value) else None
            elif info["type"] == "string":
                value = str(value).strip() if cell_has_value(value) else None

            # shorthand breed
            if metric.endswith("main_breed_variety") and cell_has_value(value):
                if value in ["HF", "hf", "Hf"]:
                    value = "Holstein"

            # grazing quality translation
            if metric.endswith("grazing_quality") and cell_has_value(value):
                grazing_quality_mapping = {
                    "wysoka": "HIGH",
                    "niska": "LOW",
                }
                value = grazing_quality_mapping.get(slugify(value), value)

            # feed conversion (NOW SAFE)
            if metric.startswith("feed."):
                value = normalize_feed_value(
                    value=value,
                    metric=metric,
                    feed_meta=feed_meta,
                    row_data=row_data,
                    dmi_conversion=dmi_conversion,
                    herd_feed_indicator=herd_feed_indicator,
                    multiday_feed_indicator=multiday_feed_indicator,
                    debug=False,
                )

            # normalize farm id
            if metric == "farm_id" and cell_has_value(value):
                value = slugify(value)

            # bedding translation
            if metric == "bedding.type" and cell_has_value(value):
                bedding_type_mapping = {
                    "sloma": "straw",
                    "soma": "straw",
                    "piasek": "sand",
                    "gazeta": "newspaper",
                    "trockenmist": "sawdust",
                    "trociny": "sawdust",
                    "inne": "newspaper"
                }
                value = bedding_type_mapping.get(slugify(value), value)

        except Exception as e:
            st.error(f"{survey.name} failed on metric {metric}: {e}")
            value = None

        row_data[metric] = value


    farm_id = row_data.get("farm_id")
    milk_year = row_data.get("milk_year")

    row_data["survey_id"] = f"{str(farm_id).strip()}_{int(milk_year)}"

    survey_loader = pd.concat(
        [survey_loader, pd.DataFrame([row_data])],
        ignore_index=True
    )



def validation_rules():
    return {
        "herd_sections": [s["display_name"] for s in herd_sections],
        "herd_varieties": [s["cft_name"] for s in herd_varieties]
    }


# -----------------------------
# Streamlit UI for Corrections
# -----------------------------
def display_error_correction_ui(error_report, df):
    if not error_report:
        st.success("✅ All data passed validation!")
        return df, True

    st.warning(f"⚠️ Found {len(error_report)} rows with data quality issues")
    st.caption(
        "Work through each issue shown below, updating values where needed. "
        "Use the buttons at the bottom to move between records or reset all changes."
    )

    # Reset state if the underlying df changes (simple + reliable)
    df_sig = tuple(df["survey_id"].astype(str).tolist())
    if st.session_state.get("dq_df_sig") != df_sig:
        st.session_state.dq_df_sig = df_sig
        st.session_state.current_error_idx = 0
        st.session_state.corrected_df = df.copy()

    current_idx = st.session_state.current_error_idx

    if current_idx >= len(error_report):
        st.success("✅ All errors reviewed!")
        return st.session_state.corrected_df, True

    current_error = error_report[current_idx]
    survey_id = current_error.get("survey_id")

    if not survey_id:
        st.error("Internal error: survey_id missing from validation error")
        st.stop()

    matches = st.session_state.corrected_df.index[
        st.session_state.corrected_df["survey_id"] == survey_id
    ]

    if len(matches) != 1:
        st.error(f"Internal error: expected 1 row for survey_id '{survey_id}', found {len(matches)}")
        st.stop()

    row_idx = matches[0]

    st.progress((current_idx + 1) / len(error_report))
    st.write(f"**Error {current_idx + 1} of {len(error_report)}**")

    identifier = current_error["row_data"].get("farm_name", survey_id)
    st.error(f"### ❌ {identifier} has {len(current_error['errors'])} error(s)")

    corrections = {}

    for col_name, error_info in current_error["errors"].items():
        with st.container():
            st.markdown(f"**Column: `{col_name}`**")

            col1, col2 = st.columns([1, 2])

            with col1:
                st.write("**Current value:**")
                st.code(str(error_info["current_value"]))
                st.write("**Issues:**")
                for err in error_info["errors"]:
                    st.write(f"- {err}")

            with col2:
                rules = error_info["rules"]

                if rules.get("type") == "categorical":
                    corrected = st.selectbox(
                        f"Fix {col_name}",
                        options=rules["allowed_values"],
                        key=f"fix_{survey_id}_{col_name}",
                        label_visibility="collapsed",
                    )
                elif rules.get("type") in ["numeric", "integer"]:
                    initial = float(error_info["current_value"]) if pd.notna(error_info["current_value"]) else 0.0
                    corrected = st.number_input(
                        f"Fix {col_name}",
                        value=initial,
                        min_value=rules.get("min"),
                        max_value=rules.get("max"),
                        key=f"fix_{survey_id}_{col_name}",
                        label_visibility="collapsed",
                    )
                else:
                    corrected = st.text_input(
                        f"Fix {col_name}",
                        value=str(error_info["current_value"]) if pd.notna(error_info["current_value"]) else "",
                        key=f"fix_{survey_id}_{col_name}",
                        label_visibility="collapsed",
                    )

                corrections[col_name] = corrected

            st.divider()

    col1, col2, col3 = st.columns([1, 1, 1])

    with col1:
        if current_idx > 0 and st.button("⬅️ Previous", use_container_width=True):
            st.session_state.current_error_idx -= 1
            st.rerun()

    with col2:
        if st.button("✅ Apply & Continue", use_container_width=True, type="primary"):
            for col_name, value in corrections.items():
                st.session_state.corrected_df.at[row_idx, col_name] = value
            st.session_state.current_error_idx += 1
            st.rerun()

    with col3:
        if st.button("🔄 Reset All", use_container_width=True):
            st.session_state.current_error_idx = 0
            st.session_state.corrected_df = df.copy()
            st.rerun()

    return st.session_state.corrected_df, False


# ----- get into columns
def flatten_cft_response(response: list) -> pd.DataFrame:
    """
    Flattens a CFT API response into a single wide table.
    One row per farm, with category-level emissions expanded into columns.
    Splits farm_identifier into farm_name and farm_year.
    """

    rows = []

    for record in response:

        farm_identifier = record["farm"]["farm_identifier"]


        if "_" in farm_identifier:
            farm_id, milk_year = farm_identifier.rsplit("_", 1)
        else:
            farm_id = farm_identifier
            milk_year = None
            st.warning("Farm identifier missing year suffix. Contact administrator.")

        summary = record["summary"]
        disagg = summary["disaggregation_totals"][0]

        row = {
            "survey_id": farm_identifier,
            "farm_id": farm_id,
            "milk_year": pd.to_numeric(milk_year, errors="coerce"),

            # Overall summary
            "emissions_total": float(summary["emissions_total"][0]),
            "emissions_total_unit": summary["emissions_total"][1],
            "emissions_per_fpcm": float(summary["emissions_per_fpcm"][0]),
            "emissions_per_fpcm_unit": summary["emissions_per_fpcm"][1],

            # Disaggregation totals
            "CO2_tonnes": float(disagg["CO2"]["metric_tonnes_CO2"][0]),
            "CO2e_from_CO2_tonnes": float(disagg["CO2"]["metric_tonnes_CO2e"][0]),

            "N2O_tonnes": float(disagg["N2O"]["metric_tonnes_N2O"][0]),
            "CO2e_from_N2O_tonnes": float(disagg["N2O"]["metric_tonnes_CO2e"][0]),

            "CH4_tonnes": float(disagg["CH4"]["metric_tonnes_CH4"][0]),
            "CO2e_from_CH4_tonnes": float(disagg["CH4"]["metric_tonnes_CO2e"][0]),

            # Metadata
            "cft_version": record["information"]["cft_version"]
        }

        # Flatten category-level emissions into wide columns
        for item in record["total_emissions"]:
            name = item["name"]

            row[f"{name}_CO2"] = float(item["CO2"])
            row[f"{name}_N2O"] = float(item["N2O"])
            row[f"{name}_CH4"] = float(item["CH4"])
            row[f"{name}_total_CO2e"] = float(item["total_CO2e"])
            row[f"{name}_total_CO2e_per_fpcm"] = float(item["total_CO2e_per_fpcm"])

        rows.append(row)

    df_wide = pd.DataFrame(rows)

    return df_wide


# -----------------------------
# Integration with your existing code
# -----------------------------
if not survey_loader.empty:
    survey_loader = survey_loader.drop_duplicates(subset=["survey_id"])
    # Step 1: Check for duplicates in database
    duplicate_rows, cleaned_df = check_duplicates_in_database(
        survey_loader, df, id_column="farm_id"
    )

    # If every uploaded survey was an exact duplicate, show info only and do not proceed
    if cleaned_df.empty and not duplicate_rows:
        st.info(
            "All uploaded survey workbook(s) are exact duplicates of existing records. "
            "No new data to validate or upload."
        )
        st.stop()

    # Show only the truly new/changed surveys
    st.write("Unique surveys", cleaned_df)

    st.caption(
        "If a farm already exists, choose whether to overwrite its data with this upload "
        "or keep the existing record."
    )
    resolved_df, duplicates_resolved = display_duplicate_resolution_ui(
        duplicate_rows, cleaned_df, df
    )
    
    # Step 2: Only proceed to validation if duplicates are resolved
    if duplicates_resolved:
        st.divider()
        st.subheader("📋 Data Quality Validation")
        st.caption(
            "Review and fix any highlighted cells below. Red warnings indicate values "
            "that must be corrected before you can submit."
        )
        
        # Run validation
        validation_rules = define_validation_rules()
        error_report = validate_dataframe(resolved_df, validation_rules)
        
        # Display correction UI
        corrected_df, all_valid = display_error_correction_ui(error_report, resolved_df)

        # Ensure numeric columns are never null
        numeric_cols = corrected_df.select_dtypes(include=["number"]).columns.tolist()
        corrected_df[numeric_cols] = corrected_df[numeric_cols].fillna(0)
        
        # Only show submit button when all valid
        if all_valid:

            # ---- derive survey_id (business identifier) ----
            if "survey_id" not in corrected_df.columns:
                corrected_df["survey_id"] = (
                    corrected_df["farm_id"].astype(str).str.strip()
                    + "_"
                    + corrected_df["milk_year"].astype(int).astype(str)
                )

            # Optional sanity check (recommended)
            if corrected_df["survey_id"].isna().any():
                st.error("Internal error: survey_id could not be derived for all rows.")
                st.stop()

            # form to submit api
            with st.form("data_quality_form"):
                st.write("✅ Data validated successfully!")
                st.caption(
                    "Submitting will upload these records to the central database and run the CFT model. "
                    "This may take up to a minute."
                )
                st.dataframe(corrected_df)
                
                if st.form_submit_button("Upload to Database and Run CFT API"):

                    # -------------------------------------------------
                    # Get overwrites from duplicate decisions
                    # -------------------------------------------------
                    overwrite_ids = []
                    if 'duplicate_decisions' in st.session_state:
                        overwrite_ids = [
                            farm_id for farm_id, decision 
                            in st.session_state.duplicate_decisions.items() 
                            if decision == "overwrite"
                        ]

                    st.info(f"Uploading {len(corrected_df)} record(s)... ({len(overwrite_ids)} overwrites)")

                    progress_bar = st.progress(0)

                    # -------------------------------------------------
                    # Build payloads
                    # -------------------------------------------------
                    records = corrected_df.to_dict(orient="records")

                    # If you want true upsert semantics for *all* rows:
                    # upsert_dairy_inputs(records)

                    # If you want overwrite to mean "hard replace":
                    new_rows = []
                    overwrite_rows = []

                    for r in records:
                        farm_id = r.get("farm_id")
                        if farm_id in overwrite_ids:
                            overwrite_rows.append(r)
                        else:
                            new_rows.append(r)

                    try:
                        # -------------------------------------------------
                        # Run CFT API first
                        # -------------------------------------------------
                        numeric_cols = corrected_df.select_dtypes(include="number").columns
                        corrected_df[numeric_cols] = corrected_df[numeric_cols].round(6)

                        api_results = submit_new_surveys(corrected_df)
                        if not api_results:
                            st.error("CFT API returned no results.")
                            st.stop()

                        # Flatten
                        df_wide = flatten_cft_response(api_results)
                        st.write(df_wide)

                        # -------------------------------------------------
                        # API succeeded — now upsert inputs
                        # -------------------------------------------------
                        if new_rows:
                            upsert_dairy_inputs(new_rows)

                        progress_bar.progress(0.5)

                        if overwrite_rows:
                            upsert_dairy_inputs(overwrite_rows)

                        progress_bar.progress(1.0)

                        stn.success(f"🎉 All {len(records)} record(s) uploaded successfully!")

                        # -------------------------------------------------
                        # Write outputs to Supabase
                        # -------------------------------------------------
                        upsert_outputs_from_df(df_wide)

                    except Exception as e:
                        st.error("❌ Failed to upload records to Supabase")
                        st.exception(e)





